import requests
from openpyxl import load_workbook

file_path = './data/financehub.xlsx'
endpoint = 'http://localhost:8080/financehub/api/v1/incomes'

wb = load_workbook(filename=file_path)
sheet = wb.active

STATUS_COLUMN_INDEX = 6

resultados = []

for row in sheet.iter_rows(min_row=2):
    reference_date = row[0].value
    gross_amount = row[1].value
    discount_amount = row[2].value if row[2].value is not None else 0.0
    net_amount = row[3].value
    income_type = row[4].value if row[4].value else "SALARY"
    description = row[5].value if row[5].value else ""
    status_cell = row[STATUS_COLUMN_INDEX]

    if status_cell.value is not None:
        continue

    if not all([reference_date, gross_amount, net_amount]):
        continue

    payload = {
        "referenceDate": str(reference_date),
        "grossAmount": float(gross_amount),
        "discountAmount": float(discount_amount),
        "netAmount": float(net_amount),
        "type": income_type,
        "description": description
    }

    try:
        response = requests.post(endpoint, json=payload)
        status_code = response.status_code
        if status_code == 200:
            status_cell.value = "OK"
            resultados.append((status_code, f"{reference_date} - {gross_amount}", "OK"))
        else:
            resultados.append((status_code, f"{reference_date} - {gross_amount}", response.text))
    except Exception as e:
        resultados.append(("ERROR", f"{reference_date} - {gross_amount}", str(e)))

wb.save(filename=file_path)

print("\n=== IMPORT REPORT ===")
for status, key, message in resultados:
    print(f"[{status}] {key} => {message}")

erros = [r for r in resultados if r[0] != 200]
if erros:
    print("\n=== ERRORS FOUND ===")
    for status, key, message in erros:
        print(f"[{status}] {key} => {message}")
else:
    print("\nAll records were imported.")
